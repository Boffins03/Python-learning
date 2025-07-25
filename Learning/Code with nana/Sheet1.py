# import django
import os 
import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Calculating the no of product per company
products_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}
product_list.cell(1,5).value = 'Inventory Price'
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_no = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

# calculating the total value inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# calculating the product with less than 10 inventory
    if inventory < 10:
        product_under_10_inv[int(product_no)] = int(inventory)

# add the value of inventory price in sheet
    inventory_price.value = inventory * price

inv_file.save('inventory2.xlsx')

print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)